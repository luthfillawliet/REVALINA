from lib2to3.pgen2 import driver
from os import link
from threading import TIMEOUT_MAX
from tkinter import E
from typing import Pattern
from matplotlib.pyplot import cla
from numpy import true_divide
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
#Library telegram
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, JobQueue
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram.update import Update
from telegram.ext.callbackcontext import CallbackContext
from pywinauto.application import Application
import requests
# import Action chains
from selenium.webdriver import ActionChains
# IMport library select
from selenium.webdriver.support.ui import Select
import time
import pyautogui

#To write into excel
from openpyxl import Workbook, load_workbook
import os

#Import file txdirectory
import txdirect

from setuptools import Command

#setting userame login
usernamesap = r"pusat\rijal.wicaksono"
passwordsap = "9214611qq?"

class InfoMaterial:
    def __init__(self,driverpath,chrome_options,scdirektori,tokenbot,chatid,pathori,pathtx,showdelay,spreadsheetdelay,downloaddelay):
        self.driverpath = driverpath
        self.scdirektori = scdirektori
        self.tokenbot = tokenbot
        self.chatid = chatid
        self.pathori = pathori
        self.pathtx = pathtx
        self.chrome_options = chrome_options
        self.showdelay = showdelay
        self.spreadsheetdelay = spreadsheetdelay
        self.downloaddelay = downloaddelay

    def getinfomaterial(self,direktori,filenamereportharian,fileexport,kodematerial,x1,y1,x2,y2):
        try:
            #Write value into excel file
            wb = load_workbook(filenamereportharian)
            ws = wb.active
            # val = ws["C1"].value
            # print(val)
            inputkode = ""
            isstring = isinstance(kodematerial,str)
            if(isstring):
                inputkode = "00000000"+kodematerial
            else:
                inputkode = "00000000"+str(kodematerial)
            print(inputkode)
            ws["C1"].value = inputkode
            print(ws["C1"].value)
            wb.save(filenamereportharian)
            time.sleep(2)
            try:
                #
                print("Membuka file report harian")
                os.startfile(filenamereportharian)
                time.sleep(5)
                #pyautogui.click(851,608)
                
                
                print("Membuka file export")
                os.startfile((fileexport))
                time.sleep(3)
                #Close data FIle Excelnya

                pyautogui.keyDown('alt')
                pyautogui.press('f4')
                pyautogui.keyUp('alt')
                #buka enkripsi
                try:
                    myscreenshoot = pyautogui.screenshot(region=(x1,y1,(x2-x1),(y2-y1)))
                    myscreenshoot.save("range.png")
                    pyautogui.keyDown('alt')
                    pyautogui.press('f4')
                    pyautogui.keyUp('alt')
                    pyautogui.press("enter")
                    return "berhasil"
                except:
                    return "gagal screenshoot"
            except:
                return "Gagal Buka excel"
        except:
            return "gagal write"

    def infokwhmeter(self,username,password):
        driver = webdriver.Chrome(self.driverpath)
        driver.get("http://erpappw1.pusat.corp.pln.co.id/irj/portal")
        print("Berhasil buka halaman SAP")
        time.sleep(1)
        try:
            usertf = WebDriverWait(driver,20).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/span/table/tbody/tr[2]/td[2]/table/tbody/tr/td[3]/table/tbody/tr[2]/td/div/form/table/tbody/tr[2]/td[2]/input"))
            )
            usertf.send_keys(usernamesap)
            password = WebDriverWait(driver,20).until(
                EC.presence_of_element_located((By.ID, "logonpassfield"))
            )
            password.send_keys(passwordsap)
            time.sleep(1)
            btnLogin =  WebDriverWait(driver,20).until(
                EC.presence_of_element_located((By.NAME, "uidPasswordLogon"))
            )
            btnLogin.click()
            time.sleep(2)
            try:
                aplikasierp = WebDriverWait(driver,20).until(
                    EC.presence_of_element_located((By.ID, "navNodeAnchor_1_1"))
                )
                aplikasierp.click()
                time.sleep(2)
                print("Berhasil direct ke Aplikasi ERP")
                # erplogistik = WebDriverWait(driver,20).until(
                #     EC.presence_of_element_located((By.XPATH, "/html/body/table/tbody/tr[1]/td/div[1]/table/tbody/tr/td[2]/div/table/tbody/tr/td/div[2]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/form/table/tbody/tr[2]/td[1]/table/tbody/tr[1]/td[2]/a"))
                # )
                # erplogistik.click()
                # print("Berhasil klik ERP Logistik")

                #Beralih untuk open tx direct
                try:
                    pathori = "C:\\xampp\\htdocs\\python\\REWA\\TX\\"
                    pathtx = "C:\\xampp\\htdocs\\python\\REWA\\"
                    getmaterial = txdirect.eksekusiSAP(pathori,pathtx)
                    getmaterial.eksekusi(118,214,281,38)
                    time.sleep(3)
                    message = "BERHASIL"
                    print(message)
                    return message
                except:
                    message = "Gagal masuk ke Aplikasi SAP Desktop"
                    print(message)
                    return message
            except:
                message = "Gagal Direct link ke SAP Desktop"
                print(message)
                return message
        except:
            message = "Gagal Login ke SAP Web, harap cek username dan password"
            print(message)
            return message

    def reportmaterial(self,storagelocation,kodematerial,bulan,tahun,filepathexcel):
        driver = webdriver.Chrome(self.driverpath,options=self.chrome_options)
        driver.get("http://erpappw1.pusat.corp.pln.co.id/irj/portal")
        driver.maximize_window()
        print("Berhasil buka halaman SAP")
        time.sleep(1)
        try:
            usertf = WebDriverWait(driver,20).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/span/table/tbody/tr[2]/td[2]/table/tbody/tr/td[3]/table/tbody/tr[2]/td/div/form/table/tbody/tr[2]/td[2]/input"))
            )
            usertf.send_keys(usernamesap)
            password = WebDriverWait(driver,20).until(
                EC.presence_of_element_located((By.ID, "logonpassfield"))
            )
            password.send_keys(passwordsap)
            time.sleep(1)
            btnLogin =  WebDriverWait(driver,20).until(
                EC.presence_of_element_located((By.NAME, "uidPasswordLogon"))
            )
            btnLogin.click()
            time.sleep(2)
            try:
                aplikasierp = WebDriverWait(driver,20).until(
                    EC.presence_of_element_located((By.ID, "navNodeAnchor_1_1"))
                )
                aplikasierp.click()
                time.sleep(2)
                print("Berhasil direct ke Aplikasi ERP")
                try:
                    getmaterial = txdirect.eksekusiSAP(self.pathori,self.pathtx,self.showdelay,self.spreadsheetdelay,self.downloaddelay)
                    message = getmaterial.eksekusireport(storagelocation,kodematerial,bulan,tahun,165,344,151,178,262,468,filepathexcel)
                    time.sleep(1)
                    print(message)
                    return message
                except:
                    message = "Gagal masuk ke Aplikasi SAP Desktop"
                    print(message)
                    return message
            except:
                message = "Gagal Direct link ke SAP Desktop"
                print(message)
                return message
        except:
            message = "Gagal Login ke SAP Web, harap cek username dan password"
            print(message)
            return message